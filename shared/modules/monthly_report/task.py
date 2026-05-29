from module.tasks.base import BaseTask, TaskResult, TaskStatus
from module.config.config import DATA_DIR
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict
import json
import traceback

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None


class MonthlyReportTask(BaseTask):
    task_id = "monthly_report"
    task_name = "运营月报"

    def run(self) -> TaskResult:
        self.status = TaskStatus.RUNNING
        start_time = datetime.now()
        self.log("INFO", "开始生成运营月报")

        try:
            month_start, month_end = self._get_report_period()
            month_label = month_start.strftime("%Y年%m月")
            self.update_progress(10, f"统计 {month_label} 数据...")

            report = {
                "month_label": month_label,
                "month_start": month_start,
                "month_end": month_end,
                "douyin": None,
                "xiaohongshu": None,
                "inspection": None,
            }

            if self.config.get("include_douyin", True):
                report["douyin"] = self._analyze_douyin(month_start, month_end)
                self.log("INFO", f"抖音本月新增 {report['douyin']['total_posts']} 条")

            self.update_progress(35, "抖音数据统计完成")

            if self.config.get("include_xiaohongshu", True):
                report["xiaohongshu"] = self._analyze_xiaohongshu(month_start, month_end)
                self.log("INFO", f"小红书本月新增 {report['xiaohongshu']['total_posts']} 条")

            self.update_progress(60, "小红书数据统计完成")

            if self.config.get("include_inspection", True):
                report["inspection"] = self._analyze_inspection(month_start, month_end)
                self.log("INFO", f"巡检本月记录 {report['inspection']['total_records']} 条")

            self.update_progress(80, "巡检数据统计完成")

            notify_title = f"📊 门店运营月报 - {month_label}"
            notify_content = self._format_notify_content(report)
            attachment_path = self._write_excel(report)

            self.status = TaskStatus.COMPLETED
            self.update_progress(100, "月报生成完成")
            self.log("SUCCESS", f"月报生成完成: {month_label}")

            return TaskResult(
                success=True,
                message=f"{month_label} 月报生成完成",
                data=self._build_result_data(report),
                start_time=start_time,
                end_time=datetime.now(),
                notify_title=notify_title,
                notify_content=notify_content,
                attachment_path=attachment_path,
            )
        except Exception as e:
            self.status = TaskStatus.ERROR
            self.log("ERROR", f"月报生成失败: {str(e)}")
            self.log("ERROR", traceback.format_exc())
            return TaskResult(
                success=False,
                message=f"月报生成失败: {str(e)}",
                error=traceback.format_exc(),
                start_time=start_time,
                end_time=datetime.now(),
            )

    def _get_report_period(self):
        report_month = (self.config.get("report_month") or "").strip()
        if report_month:
            month_start = datetime.strptime(report_month, "%Y-%m")
        else:
            today = datetime.now()
            first_day = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            last_month_last_day = first_day - timedelta(days=1)
            month_start = last_month_last_day.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1)
        return month_start, month_end

    def _analyze_douyin(self, month_start: datetime, month_end: datetime):
        data_file = DATA_DIR / "douyin" / "posts.json"
        posts = self._load_posts(data_file, "posts")
        month_posts = [p for p in posts if self._in_month(p.get("crawl_time"), month_start, month_end)]
        return self._analyze_posts(month_posts, self._get_douyin_store_names(), include_link=True)

    def _analyze_xiaohongshu(self, month_start: datetime, month_end: datetime):
        data_file = DATA_DIR / "xiaohongshu_data.json"
        posts = self._load_posts(data_file, "posts")
        month_posts = [p for p in posts if self._in_month(p.get("crawl_time"), month_start, month_end)]
        return self._analyze_posts(month_posts, self._get_xiaohongshu_store_names(), include_link=False)

    def _analyze_inspection(self, month_start: datetime, month_end: datetime):
        data_file = DATA_DIR / "inspection_data.json"
        records = self._load_posts(data_file, "stores")
        month_records = [r for r in records if self._in_month(r.get("crawl_time"), month_start, month_end)]
        latest_by_store = {}
        for record in month_records:
            name = record.get("short_name") or record.get("name") or "未知门店"
            if name not in latest_by_store or record.get("crawl_time", "") > latest_by_store[name].get("crawl_time", ""):
                latest_by_store[name] = record

        scores = [self._safe_float(r.get("monthly_score")) for r in latest_by_store.values()]
        counts = [self._safe_int(r.get("monthly_count")) for r in latest_by_store.values()]
        avg_score = round(sum(scores) / len(scores), 2) if scores else 0
        full_score_count = sum(1 for score in scores if score >= 100)
        low_score = sorted(
            [
                {
                    "name": name,
                    "monthly_count": self._safe_int(record.get("monthly_count")),
                    "monthly_score": self._safe_float(record.get("monthly_score")),
                    "yearly_score": self._safe_float(record.get("yearly_score")),
                }
                for name, record in latest_by_store.items()
                if self._safe_float(record.get("monthly_score")) < 100
            ],
            key=lambda item: item["monthly_score"]
        )

        ranking = sorted(
            [
                {
                    "name": name,
                    "monthly_count": self._safe_int(record.get("monthly_count")),
                    "monthly_score": self._safe_float(record.get("monthly_score")),
                    "yearly_score": self._safe_float(record.get("yearly_score")),
                }
                for name, record in latest_by_store.items()
            ],
            key=lambda item: (-item["monthly_score"], -item["monthly_count"], item["name"])
        )

        return {
            "total_records": len(month_records),
            "store_count": len(latest_by_store),
            "avg_score": avg_score,
            "full_score_count": full_score_count,
            "avg_monthly_count": round(sum(counts) / len(counts), 2) if counts else 0,
            "low_score": low_score,
            "ranking": ranking,
            "records": month_records,
        }

    def _analyze_posts(self, posts: list, configured_stores: list, include_link: bool):
        top_limit = self._safe_int(self.config.get("top_limit")) or 5
        by_store = defaultdict(lambda: {"count": 0, "likes": 0, "posts": []})

        for post in posts:
            store_name = post.get("store_name") or "未知门店"
            likes = self._safe_int(post.get("likes"))
            by_store[store_name]["count"] += 1
            by_store[store_name]["likes"] += likes
            by_store[store_name]["posts"].append(post)

        ranking = []
        for store_name, stats in by_store.items():
            count = stats["count"]
            ranking.append({
                "name": store_name,
                "count": count,
                "likes": stats["likes"],
                "avg_likes": round(stats["likes"] / count, 2) if count else 0,
            })
        ranking.sort(key=lambda item: (-item["count"], -item["likes"], item["name"]))

        zero_stores = []
        if self.config.get("zero_post_warning", True):
            active_stores = set(by_store.keys())
            zero_stores = [name for name in configured_stores if name and name not in active_stores]

        return {
            "total_posts": len(posts),
            "total_likes": sum(self._safe_int(post.get("likes")) for post in posts),
            "active_store_count": len(by_store),
            "avg_likes": round(sum(self._safe_int(post.get("likes")) for post in posts) / len(posts), 2) if posts else 0,
            "ranking": ranking,
            "top": ranking[:top_limit],
            "zero_stores": zero_stores,
            "posts": posts,
            "include_link": include_link,
        }

    def _load_posts(self, data_file: Path, key: str):
        if not data_file.exists():
            self.log("WARNING", f"数据文件不存在: {data_file}")
            return []
        with open(data_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get(key, [])

    def _get_douyin_store_names(self):
        config_file = DATA_DIR.parent / "modules" / "douyin" / "config.json"
        data = self._load_json(config_file)
        names = []
        for account in data.get("accounts", []):
            name = account.get("name") or account.get("short_name")
            if name:
                names.append(name)
        return names

    def _get_xiaohongshu_store_names(self):
        config_file = DATA_DIR.parent / "modules" / "xiaohongshu" / "config.json"
        data = self._load_json(config_file)
        return [name for name in data.get("stores", []) if name]

    def _load_json(self, file_path: Path):
        if not file_path.exists():
            return {}
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _format_notify_content(self, report: dict):
        lines = [f"**{report['month_label']}运营月报**"]

        douyin = report.get("douyin")
        if douyin:
            lines.extend(self._format_platform_summary("抖音", douyin))

        xiaohongshu = report.get("xiaohongshu")
        if xiaohongshu:
            lines.extend(self._format_platform_summary("小红书", xiaohongshu))

        inspection = report.get("inspection")
        if inspection:
            lines.extend([
                "",
                "**巡检评分**",
                f"> 覆盖门店：{inspection['store_count']}家 | 平均分：{inspection['avg_score']}",
                f"> 满分门店：{inspection['full_score_count']}家 | 月均次数：{inspection['avg_monthly_count']}",
            ])
            if inspection["low_score"]:
                lines.append(f"> 需关注：{'、'.join(item['name'] for item in inspection['low_score'][:5])}")

        lines.extend(["", "_详细数据见邮件附件_", f"_{datetime.now().strftime('%m-%d %H:%M')}_"])
        return "\n".join(lines)

    def _format_platform_summary(self, platform_name: str, stats: dict):
        lines = [
            "",
            f"**{platform_name}平台**",
            f"> 新增：{stats['total_posts']}条 | 获赞：{stats['total_likes']} | 活跃门店：{stats['active_store_count']}家",
        ]
        if stats["top"]:
            lines.append("> TOP门店：")
            for index, item in enumerate(stats["top"], 1):
                lines.append(f"> {index}. {item['name']} - {item['count']}条 / {item['likes']}赞")
        if stats["zero_stores"]:
            shown = "、".join(stats["zero_stores"][:8])
            extra = f" 等{len(stats['zero_stores'])}家" if len(stats["zero_stores"]) > 8 else ""
            lines.append(f"> 零发布：{shown}{extra}")
        return lines

    def _write_excel(self, report: dict):
        if Workbook is None:
            self.log("WARNING", "未安装 openpyxl，跳过 Excel 附件生成")
            return None

        report_dir = DATA_DIR / "monthly_report"
        report_dir.mkdir(parents=True, exist_ok=True)
        month_key = report["month_start"].strftime("%Y%m")
        file_path = report_dir / f"monthly_report_{month_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "概览"
        self._write_overview_sheet(ws, report)

        if report.get("douyin"):
            self._write_posts_sheet(wb, "抖音明细", report["douyin"]["posts"], include_link=True)
            self._write_ranking_sheet(wb, "抖音排行", report["douyin"]["ranking"])

        if report.get("xiaohongshu"):
            self._write_posts_sheet(wb, "小红书明细", report["xiaohongshu"]["posts"], include_link=False)
            self._write_ranking_sheet(wb, "小红书排行", report["xiaohongshu"]["ranking"])

        if report.get("inspection"):
            self._write_inspection_sheet(wb, report["inspection"])

        wb.save(file_path)
        self.log("INFO", f"Excel 附件已生成: {file_path}")
        return str(file_path)

    def _write_overview_sheet(self, ws, report: dict):
        rows = [["门店运营月报", report["month_label"]], []]
        if report.get("douyin"):
            rows.append(["抖音新增", report["douyin"]["total_posts"], "抖音获赞", report["douyin"]["total_likes"]])
        if report.get("xiaohongshu"):
            rows.append(["小红书新增", report["xiaohongshu"]["total_posts"], "小红书获赞", report["xiaohongshu"]["total_likes"]])
        if report.get("inspection"):
            rows.append(["巡检覆盖门店", report["inspection"]["store_count"], "巡检平均分", report["inspection"]["avg_score"]])
        rows.append([])
        rows.append(["平台", "门店", "发布量", "点赞数", "平均点赞"])

        for platform_name, key in [("抖音", "douyin"), ("小红书", "xiaohongshu")]:
            stats = report.get(key)
            if not stats:
                continue
            for item in stats["top"]:
                rows.append([platform_name, item["name"], item["count"], item["likes"], item["avg_likes"]])

        for row in rows:
            ws.append(row)
        self._style_sheet(ws)

    def _write_posts_sheet(self, wb, title: str, posts: list, include_link: bool):
        ws = wb.create_sheet(title)
        headers = ["门店名称", "标题", "点赞数", "发布时间", "采集时间"]
        if include_link:
            headers.insert(3, "内容链接")
        ws.append(headers)
        for post in posts:
            row = [
                post.get("store_name", ""),
                post.get("title", ""),
                self._safe_int(post.get("likes")),
            ]
            if include_link:
                row.append(post.get("post_link", ""))
            row.extend([post.get("post_time", ""), post.get("crawl_time", "")])
            ws.append(row)
        self._style_sheet(ws)

    def _write_ranking_sheet(self, wb, title: str, ranking: list):
        ws = wb.create_sheet(title)
        ws.append(["排名", "门店", "发布量", "点赞数", "平均点赞"])
        for index, item in enumerate(ranking, 1):
            ws.append([index, item["name"], item["count"], item["likes"], item["avg_likes"]])
        self._style_sheet(ws)

    def _write_inspection_sheet(self, wb, inspection: dict):
        ws = wb.create_sheet("巡检评分")
        ws.append(["门店", "月度次数", "月度评分", "年度评分"])
        for item in inspection["ranking"]:
            ws.append([item["name"], item["monthly_count"], item["monthly_score"], item["yearly_score"]])
        self._style_sheet(ws)

    def _style_sheet(self, ws):
        header_fill = PatternFill("solid", fgColor="E8F4FF")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 50)

    def _build_result_data(self, report: dict):
        return {
            "month": report["month_start"].strftime("%Y-%m"),
            "douyin_new": report["douyin"]["total_posts"] if report.get("douyin") else 0,
            "xiaohongshu_new": report["xiaohongshu"]["total_posts"] if report.get("xiaohongshu") else 0,
            "inspection_stores": report["inspection"]["store_count"] if report.get("inspection") else 0,
        }

    def _in_month(self, value: str, month_start: datetime, month_end: datetime):
        parsed = self._parse_datetime(value)
        return bool(parsed and month_start <= parsed < month_end)

    def _parse_datetime(self, value: str):
        if not value:
            return None
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return None

    def _safe_int(self, value):
        try:
            return int(value or 0)
        except (TypeError, ValueError):
            return 0

    def _safe_float(self, value):
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0
